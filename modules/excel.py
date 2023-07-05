import openpyxl
import datetime
import os


class Excel:
    def __init__(self) -> None:
        self.__path = 'D:\\python\\HackMd_track_machine\\data\\record.xlsx'
        self.__specify_excel()

    def __specify_excel(self):
        if not os.path.exists(self.__path):
            self.wb = openpyxl.Workbook()
        else:
            self.wb = openpyxl.load_workbook(self.__path)
        self.sheet = self.wb[self.wb.sheetnames[0]]

    def write_record(self, spiders: object):
        this_row = self.sheet.max_row + 1  # for record what row this time log
        # write timestamp
        self.sheet[f'A{this_row}'].value = datetime.datetime.now(
        ).strftime("%Y/%m/%d %H:%M:%S")
        self.sheet.column_dimensions['A'].width = 26
        # write timestamp

        # write data
        for number, spider in enumerate(spiders, 1):
            title = spider.get_title()
            eye = spider.get_eye()
            heart = spider.get_heart()

            if self.sheet[f'{chr(65+number)}1'].value == None:
                # write title
                self.sheet[f'{chr(65+number)}1'].value = title

            elif self.sheet[f'{chr(65+number)}1'].value == "RefNum":
                self.sheet.insert_cols(number+1)
                self.sheet[f'{chr(65+number)}1'].value = title
            self.sheet[f'{chr(65+number)}{this_row}'].value = f'{eye}/{heart}'
            self.sheet.column_dimensions[f'{chr(65+number)}'].width = 26
        # write data
        if self.sheet[f'{chr(65+number+1)}1'].value == None:
            self.sheet[f'{chr(65+number+1)}1'].value = "RefNum"
        refine_number_past = self.sheet[f'{chr(65+number+1)}{this_row-1}'].value

        if type(refine_number_past) != int:
            self.sheet[f'{chr(65+number+1)}{this_row}'] = 1
        else:
            self.sheet[f'{chr(65+number+1)}{this_row}'] = int(refine_number_past+1)
        self.wb.save(self.__path)
